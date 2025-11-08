from openpyxl import load_workbook
import os


def process_excel(input_path, output_path, translator=None):
    """
    translator(text) → returns translated text.
    For micro-pilot, translator can simply return text or dummy output.
    """
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(input_path)
    ws = wb.active

    start_row = 2  # assuming row 1 is header

    for row in range(start_row, ws.max_row + 1):
        source = ws[f"D{row}"].value

        if not source:
            continue

        # translation step
        if translator:
            translated = translator(source)
        else:
            translated = source  # placeholder (pilot allowed)

        ws[f"E{row}"] = translated  # write translation into Column E

    wb.save(output_path)
    print(f"Excel processed → {output_path}")
