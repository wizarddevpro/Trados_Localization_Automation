import json
from openpyxl import load_workbook
import os


def excel_mapping(input_path, json_path):
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(json_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb = load_workbook(input_path)
    ws = wb.active

    mapping = []

    for row in range(2, ws.max_row + 1):
        source = ws[f"D{row}"].value
        translation = ws[f"E{row}"].value

        if not source:
            continue

        seg = {
            "seg_id": f"xls_r{row}",
            "row": row,
            "source": source,
            "translation": translation,
        }
        mapping.append(seg)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"segments": mapping}, f, indent=4)

    print(f"JSON mapping written → {json_path}")
