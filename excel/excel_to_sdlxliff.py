from openpyxl import load_workbook
from lxml import etree
import nltk
import os

SDL_NS = "http://sdl.com/FileTypes/SdlXliff/1.0"
NSMAP = {"sdl": SDL_NS}


def excel_to_sdlxliff(input_excel, output_sdlxliff):
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(output_sdlxliff)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb = load_workbook(input_excel)
    ws = wb.active

    root = etree.Element("xliff", version="1.2", nsmap=NSMAP)
    file_el = etree.SubElement(
        root,
        "file",
        original=input_excel,
        datatype="x-sdlfilterframework2",
        **{"source-language": "en-CA", "target-language": "fr-CA"},
    )
    body = etree.SubElement(file_el, "body")

    for row in range(2, ws.max_row + 1):
        source = ws[f"D{row}"].value
        if not source:
            continue

        seg_id = f"xls_r{row}"

        tu = etree.SubElement(body, "trans-unit", id=seg_id, translate="yes")

        src = etree.SubElement(tu, "source")
        src.text = source

    etree.ElementTree(root).write(
        output_sdlxliff,
        pretty_print=True,
        encoding="utf-8",
        xml_declaration=True,
    )

    print(f"SDLXLIFF written → {output_sdlxliff}")
